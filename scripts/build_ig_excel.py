import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = Path(
    r"C:\Users\MasterUnion\.cursor\projects\c-Users-MasterUnion-Downloads-thumbnail-generator\agent-tools\2075773f-b944-4b3d-99c4-d8d3c0139646.txt"
)
OUT = Path(
    r"C:\Users\MasterUnion\Downloads\thumbnail-generator\MU_Instagram_Videos_50k_Plus.xlsx"
)

CHANNELS = [
    "elevatorpitch.mu",
    "lifeatmu",
    "bharat.mu_",
    "builders.mu",
    "pov__mu",
    "offcampus.mu",
    "u18.club",
    "masters.union",
]


def grab(key: str, block: str):
    m = re.search(rf"(?:^|\n)\s*{re.escape(key)}:\s*(.*)", block)
    if not m:
        return None
    val = m.group(1).strip()
    if val == "null":
        return None
    if len(val) >= 2 and val[0] == val[-1] and val[0] in "\"'":
        val = val[1:-1]
    return val


def to_int(val):
    if val in (None, ""):
        return None
    try:
        return int(float(val))
    except Exception:
        return None


def to_float(val):
    if val in (None, ""):
        return None
    try:
        return float(val)
    except Exception:
        return None


def source_from_input(input_url: str | None, owner: str | None) -> str:
    if input_url:
        path = urlparse(input_url).path.strip("/")
        # https://www.instagram.com/masters.union/ or username string
        if path:
            first = path.split("/")[0]
            if first and first not in {"p", "reel", "reels"}:
                return first
        # bare username
        if "/" not in input_url and "instagram.com" not in input_url:
            return input_url.strip("@")
    return owner or ""


def main():
    text = SRC.read_text(encoding="utf-8")
    parts = re.split(r"\n  - ownerUsername:\s*", text)
    items = []
    scraped = 0

    for part in parts[1:]:
        scraped += 1
        block = "ownerUsername: " + part
        play_n = to_int(grab("videoPlayCount", block)) or 0
        view_n = to_int(grab("videoViewCount", block))
        metric = play_n if play_n else (view_n or 0)
        if metric < 50000:
            continue

        owner = grab("ownerUsername", block) or ""
        source = source_from_input(grab("inputUrl", block), owner)
        caption = (grab("caption", block) or "").replace("\\n", "\n")

        items.append(
            {
                "source_channel": source,
                "owner": owner,
                "url": grab("url", block) or "",
                "plays": play_n,
                "views": view_n if view_n is not None else "",
                "likes": to_int(grab("likesCount", block)) or "",
                "comments": to_int(grab("commentsCount", block)) or "",
                "caption": caption[:500],
                "posted_at": grab("timestamp", block) or "",
                "duration_sec": to_float(grab("videoDuration", block)) or "",
                "shortcode": grab("shortCode", block) or "",
                "type": grab("productType", block) or "",
            }
        )

    items.sort(key=lambda x: x["plays"], reverse=True)
    print(f"scraped={scraped} filtered={len(items)}")
    print("by source:", dict(Counter(i["source_channel"] for i in items)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Videos 50k+"

    headers = [
        "Source Channel",
        "Owner Username",
        "Video URL",
        "Play Count",
        "View Count",
        "Likes",
        "Comments",
        "Duration (sec)",
        "Posted At (UTC)",
        "Shortcode",
        "Product Type",
        "Caption",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for r, row in enumerate(items, 2):
        vals = [
            row["source_channel"],
            row["owner"],
            row["url"],
            row["plays"],
            row["views"],
            row["likes"],
            row["comments"],
            row["duration_sec"],
            row["posted_at"],
            row["shortcode"],
            row["type"],
            row["caption"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 12))
            if c == 3 and isinstance(v, str) and v.startswith("http"):
                cell.hyperlink = v
                cell.font = Font(color="0563C1", underline="single")

    widths = [18, 22, 48, 14, 14, 12, 12, 14, 22, 14, 14, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:L{max(1, len(items) + 1)}"
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Masters Union Instagram — Videos with >50k plays"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Scraped at (UTC)"
    ws2["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    ws2["A4"] = "Total reels scraped"
    ws2["B4"] = scraped
    ws2["A5"] = "Videos >50k plays"
    ws2["B5"] = len(items)

    ws2["A7"] = "Source Channel"
    ws2["B7"] = "Count >50k"
    ws2["C7"] = "Top play count"
    for cell in (ws2["A7"], ws2["B7"], ws2["C7"]):
        cell.fill = header_fill
        cell.font = header_font

    by = defaultdict(list)
    for i in items:
        by[i["source_channel"]].append(i["plays"])

    r = 8
    for ch in CHANNELS:
        plays = by.get(ch, [])
        ws2.cell(r, 1, ch)
        ws2.cell(r, 2, len(plays))
        ws2.cell(r, 3, max(plays) if plays else 0)
        r += 1

    for ch in sorted(by.keys()):
        if ch not in CHANNELS:
            plays = by[ch]
            ws2.cell(r, 1, ch)
            ws2.cell(r, 2, len(plays))
            ws2.cell(r, 3, max(plays) if plays else 0)
            r += 1

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 16
    ws2[f"A{r + 1}"] = "Notes"
    ws2[f"A{r + 1}"].font = Font(bold=True)
    note = (
        "Filter: Play Count > 50,000. "
        "Source Channel = profile scraped from. "
        "Owner Username may differ for collab/coauthored reels. "
        "Up to 100 latest reels per channel were scraped."
    )
    ws2[f"A{r + 2}"] = note
    ws2.merge_cells(start_row=r + 2, start_column=1, end_row=r + 2, end_column=3)
    ws2.cell(r + 2, 1).alignment = Alignment(wrap_text=True)
    ws2.row_dimensions[r + 2].height = 50

    wb.save(OUT)
    print(f"saved={OUT}")


if __name__ == "__main__":
    main()
